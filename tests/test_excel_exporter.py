"""Tests for Excel workbook exports."""

from openpyxl import load_workbook

from dataset_finder.batch import BatchSearchIssue, BatchSearchResult
from dataset_finder.exporters.excel_exporter import export_excel
from dataset_finder.models import DatasetRecord


def test_export_excel_creates_expected_sheets(tmp_path) -> None:
    record = DatasetRecord(
        uid="123",
        accession="GSE123",
        title="Drosophila brain RNA-seq",
        organism="Drosophila melanogaster",
        study_type="RNA-seq",
        sample_count=6,
        publication_date="2026-01-01",
        url="https://example.org/GSE123",
        database="GEO",
        gene="bru1",
        gene_set="RBP",
        technique="RNA_seq",
    )

    result = BatchSearchResult(
        records=(record,),
        issues=(
            BatchSearchIssue(
                gene="bad",
                database="geo",
                message="Test error",
            ),
        ),
        database_statuses=(),
        gene_annotations=(),
        genes=("bru1", "bad"),
        gene_set="RBP",
        database="geo",
    )

    output_path = export_excel(
        result,
        tmp_path / "screening.xlsx",
    )

    workbook = load_workbook(
        output_path,
        read_only=True,
    )

    assert "README" in workbook.sheetnames
    assert "Gene_Summary" in workbook.sheetnames
    assert "Gene_Annotations" in workbook.sheetnames
    assert "All_Datasets" in workbook.sheetnames

    all_datasets_headers = [
        cell.value
        for cell in workbook["All_Datasets"][1]
    ]

    for column_name in (
        "Tissue",
        "Cell Type",
        "Developmental Stage",
        "Sex",
        "Genotype",
        "Strain",
        "Treatment",
        "Control Status",
        "Disease",
        "Time Point",
        "Perturbation",
    ):
        assert column_name in all_datasets_headers
    assert "RNA_seq" in workbook.sheetnames
    assert "Errors" in workbook.sheetnames

    all_datasets = workbook["All_Datasets"]
    headers = [
        cell.value
        for cell in next(all_datasets.iter_rows())
    ]

    assert "Gene" in headers
    assert "Technique" in headers
    assert "Dataset URL" in headers


def test_neural_highlight_formula_uses_biological_columns() -> None:
    from dataset_finder.exporters.excel_exporter import (
        _neural_highlight_formula,
    )

    formula = _neural_highlight_formula(
        [
            "Gene",
            "Tissue",
            "Cell Type",
            "Title",
            "Description",
        ]
    )

    assert formula.startswith("=OR(")
    assert 'SEARCH("brain",$B2)' in formula
    assert 'SEARCH("neuron",$C2)' in formula
    assert 'SEARCH("mushroom body",$D2)' in formula
    assert 'SEARCH("glial",$E2)' in formula


def test_neural_highlight_formula_ignores_head_alone() -> None:
    from dataset_finder.exporters.excel_exporter import (
        _neural_highlight_formula,
    )

    formula = _neural_highlight_formula(
        ["Gene", "Tissue", "Title"]
    )

    assert 'SEARCH("head"' not in formula


def test_neural_highlight_formula_requires_searchable_columns() -> None:
    from dataset_finder.exporters.excel_exporter import (
        _neural_highlight_formula,
    )

    assert (
        _neural_highlight_formula(
            ["Gene", "Accession", "Database"]
        )
        == ""
    )
